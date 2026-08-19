"""Rapor üretici / Report builder: PDF, Excel, CSV.

Her rapor bir `ReportPreview` (sütunlar + satırlar + toplamlar) üretir; dışa
aktarma katmanı bu ortak yapıyı biçimlendirir. Böylece yeni rapor eklemek
yalnızca bir veri fonksiyonu yazmayı gerektirir.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.logging_config import get_logger
from app.models.attendance import Attendance
from app.models.competition import Competition
from app.models.enums import MembershipStatus, PaymentStatus, StudentStatus
from app.models.finance import Expense, Invoice, Payment
from app.models.lesson import Lesson
from app.models.membership import Membership
from app.models.people import Student
from app.models.system import AppSetting
from app.schemas.statistics import ReportDefinition, ReportPreview, ReportRequest
from app.services.formatting import (
    format_currency,
    format_date,
    format_number,
    format_percent,
    format_swim_time,
)
from app.services.statistics_engine import (
    PRESENT_STATUSES,
    attendance_statistics,
    instructor_statistics,
    pool_statistics,
    resolve_period,
    student_performance_summary,
    student_statistics,
)

logger = get_logger("application")

# ---------------------------------------------------------------------------
# Türkçe karakter desteği için font kaydı
# ---------------------------------------------------------------------------
_FONT_NAME = "Helvetica"
_FONT_BOLD = "Helvetica-Bold"


def _register_fonts() -> None:
    """Windows'ta DejaVu/Arial yüklüyse Türkçe karakterler için kaydeder."""
    global _FONT_NAME, _FONT_BOLD
    candidates = [
        (
            "DejaVuSans",
            "C:/Windows/Fonts/DejaVuSans.ttf",
            "C:/Windows/Fonts/DejaVuSans-Bold.ttf",
        ),
        ("Arial", "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("Calibri", "C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
    ]
    for name, regular, bold in candidates:
        try:
            pdfmetrics.registerFont(TTFont(name, regular))
            pdfmetrics.registerFont(TTFont(f"{name}-Bold", bold))
            _FONT_NAME, _FONT_BOLD = name, f"{name}-Bold"
            return
        except Exception:  # noqa: BLE001 - font yoksa bir sonrakini dene
            continue
    logger.warning(
        "Türkçe karakter destekli font bulunamadı; PDF'lerde Helvetica kullanılacak."
    )


_register_fonts()


def _org_name(db: Session) -> str:
    setting = db.scalar(select(AppSetting).where(AppSetting.key == "organization"))
    if setting and isinstance(setting.value, dict):
        return setting.value.get("name") or settings.app_name
    return settings.app_name


# ===========================================================================
# Rapor tanımları
# ===========================================================================
REPORT_DEFINITIONS: list[ReportDefinition] = [
    ReportDefinition(
        key="daily_manager",
        title_tr="Günlük Yönetici Raporu",
        title_en="Daily Manager Report",
        description_tr="Günün dersleri, yoklama ve tahsilat özeti",
        description_en="Today's lessons, attendance and collections",
        category="management",
        filters=["date"],
        required_permission="report:read",
    ),
    ReportDefinition(
        key="weekly_management",
        title_tr="Haftalık Yönetim Raporu",
        title_en="Weekly Management Report",
        description_tr="Haftalık operasyon ve finans özeti",
        description_en="Weekly operations and finance summary",
        category="management",
        filters=["period"],
        required_permission="report:read",
    ),
    ReportDefinition(
        key="monthly_management",
        title_tr="Aylık Yönetim Raporu",
        title_en="Monthly Management Report",
        description_tr="Aylık KPI, gelir ve öğrenci hareketleri",
        description_en="Monthly KPIs, revenue and student movement",
        category="management",
        filters=["period"],
        required_permission="report:read",
    ),
    ReportDefinition(
        key="student_list",
        title_tr="Öğrenci Listesi",
        title_en="Student List",
        description_tr="Filtrelenebilir öğrenci kayıt listesi",
        description_en="Filterable student roster",
        category="students",
        filters=["group", "status", "level"],
        required_permission="student:read",
    ),
    ReportDefinition(
        key="student_progress",
        title_tr="Öğrenci Gelişim Raporu",
        title_en="Student Progress Report",
        description_tr="Tek öğrencinin performans ve devam gelişimi",
        description_en="Single student performance and attendance progress",
        category="students",
        filters=["student", "period"],
        required_permission="performance:read",
    ),
    ReportDefinition(
        key="attendance",
        title_tr="Yoklama Raporu",
        title_en="Attendance Report",
        description_tr="Devam oranları ve devamsızlık analizi",
        description_en="Attendance rates and absence analysis",
        category="operations",
        filters=["period", "group", "instructor"],
        required_permission="attendance:read",
    ),
    ReportDefinition(
        key="instructor_workload",
        title_tr="Eğitmen İş Yükü Raporu",
        title_en="Instructor Workload Report",
        description_tr="Eğitmen başına ders, saat ve doluluk",
        description_en="Lessons, hours and occupancy per instructor",
        category="staff",
        filters=["period"],
        required_permission="instructor:read",
    ),
    ReportDefinition(
        key="pool_usage",
        title_tr="Havuz Kullanım Raporu",
        title_en="Pool Usage Report",
        description_tr="Havuz ve saat bazlı doluluk",
        description_en="Pool and hourly occupancy",
        category="facility",
        filters=["period", "pool"],
        required_permission="pool:read",
    ),
    ReportDefinition(
        key="lane_occupancy",
        title_tr="Kulvar Doluluk Raporu",
        title_en="Lane Occupancy Report",
        description_tr="Kulvar bazlı kullanım dağılımı",
        description_en="Per-lane utilisation breakdown",
        category="facility",
        filters=["period", "pool"],
        required_permission="pool:read",
    ),
    ReportDefinition(
        key="finance",
        title_tr="Finans Raporu",
        title_en="Finance Report",
        description_tr="Gelir, gider ve net kâr",
        description_en="Income, expenses and net profit",
        category="finance",
        filters=["period"],
        required_permission="finance:read",
    ),
    ReportDefinition(
        key="collections",
        title_tr="Tahsilat Raporu",
        title_en="Collections Report",
        description_tr="Tahsilatlar ve yöntem dağılımı",
        description_en="Collections and payment method breakdown",
        category="finance",
        filters=["period"],
        required_permission="finance:read",
    ),
    ReportDefinition(
        key="outstanding",
        title_tr="Bekleyen Alacak Raporu",
        title_en="Outstanding Receivables",
        description_tr="Geciken ve bekleyen ödemeler (yaşlandırma)",
        description_en="Overdue and pending payments (aging)",
        category="finance",
        filters=[],
        required_permission="finance:read",
    ),
    ReportDefinition(
        key="membership",
        title_tr="Üyelik Raporu",
        title_en="Membership Report",
        description_tr="Aktif, biten ve dondurulmuş üyelikler",
        description_en="Active, expired and frozen memberships",
        category="sales",
        filters=["status"],
        required_permission="membership:read",
    ),
    ReportDefinition(
        key="sales",
        title_tr="Satış Raporu",
        title_en="Sales Report",
        description_tr="Paket bazlı satış ve gelir",
        description_en="Package-level sales and revenue",
        category="sales",
        filters=["period"],
        required_permission="finance:read",
    ),
    ReportDefinition(
        key="performance",
        title_tr="Performans Raporu",
        title_en="Performance Report",
        description_tr="Sporcu dereceleri ve gelişim",
        description_en="Athlete times and improvement",
        category="sports",
        filters=["period", "student"],
        required_permission="performance:read",
    ),
    ReportDefinition(
        key="competition",
        title_tr="Yarışma Raporu",
        title_en="Competition Report",
        description_tr="Yarışma sonuçları ve madalyalar",
        description_en="Competition results and medals",
        category="sports",
        filters=["period"],
        required_permission="competition:read",
    ),
]

REPORT_INDEX = {definition.key: definition for definition in REPORT_DEFINITIONS}


# ===========================================================================
# Rapor veri üreticileri
# ===========================================================================
def _col(key: str, tr: str, en: str, lang: str) -> dict[str, str]:
    return {"key": key, "label": tr if lang == "tr" else en}


def _report_student_list(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    stmt = select(Student)
    if request.group_id:
        stmt = stmt.where(Student.group_id == request.group_id)
    if request.instructor_id:
        stmt = stmt.where(Student.primary_instructor_id == request.instructor_id)
    students = db.scalars(stmt.order_by(Student.last_name, Student.first_name)).all()

    rows = [
        {
            "student_number": s.student_number,
            "full_name": s.full_name,
            "age": s.age or "-",
            "swim_level": s.swim_level,
            "status": s.status,
            "group": s.group.name if s.group else "-",
            "instructor": (
                s.primary_instructor.full_name if s.primary_instructor else "-"
            ),
            "registration_date": format_date(s.registration_date, lang),
            "phone": s.phone or (s.guardians[0].guardian.phone if s.guardians else "-"),
        }
        for s in students
    ]

    return ReportPreview(
        report_key="student_list",
        title="Öğrenci Listesi" if lang == "tr" else "Student List",
        generated_at=datetime.now(),
        period_label="-",
        columns=[
            _col("student_number", "Öğrenci No", "Student No", lang),
            _col("full_name", "Ad Soyad", "Full Name", lang),
            _col("age", "Yaş", "Age", lang),
            _col("swim_level", "Seviye", "Level", lang),
            _col("status", "Durum", "Status", lang),
            _col("group", "Grup", "Group", lang),
            _col("instructor", "Eğitmen", "Instructor", lang),
            _col("registration_date", "Kayıt Tarihi", "Registered", lang),
            _col("phone", "Telefon", "Phone", lang),
        ],
        rows=rows,
        totals={"count": len(rows)},
        summary={
            "total": len(rows),
            "active": sum(1 for s in students if s.status == StudentStatus.ACTIVE),
        },
        row_count=len(rows),
    )


def _report_attendance(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    stats = attendance_statistics(
        db, start, end, request.group_id, request.instructor_id
    )

    rows = [
        {
            "metric": (
                "Genel Devam Oranı" if lang == "tr" else "Overall Attendance Rate"
            ),
            "value": format_percent(stats.overall_rate, lang),
        },
        {
            "metric": "Geldi" if lang == "tr" else "Present",
            "value": stats.present_count,
        },
        {
            "metric": "Gelmedi" if lang == "tr" else "Absent",
            "value": stats.absent_count,
        },
        {"metric": "Geç Geldi" if lang == "tr" else "Late", "value": stats.late_count},
        {
            "metric": "Mazeretli" if lang == "tr" else "Excused",
            "value": stats.excused_count,
        },
        {
            "metric": "Telafi" if lang == "tr" else "Make-up",
            "value": stats.makeup_count,
        },
        {
            "metric": "Devamsızlık Oranı" if lang == "tr" else "No-show Rate",
            "value": format_percent(stats.no_show_rate, lang),
        },
    ]
    rows.extend(
        {
            "metric": f"{'Grup' if lang == 'tr' else 'Group'}: {d.label}",
            "value": format_percent(d.value, lang),
        }
        for d in stats.by_group
    )

    return ReportPreview(
        report_key="attendance",
        title="Yoklama Raporu" if lang == "tr" else "Attendance Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("metric", "Gösterge", "Metric", lang),
            _col("value", "Değer", "Value", lang),
        ],
        rows=rows,
        totals={},
        summary=stats.model_dump(mode="json"),
        row_count=len(rows),
    )


def _report_instructor_workload(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    stats = instructor_statistics(db, start, end)

    rows = [
        {
            "full_name": r.full_name,
            "lesson_count": r.lesson_count,
            "total_hours": format_number(r.total_hours, lang, 1),
            "student_count": r.student_count,
            "occupancy_rate": format_percent(r.occupancy_rate, lang),
            "attendance_rate": format_percent(r.attendance_rate, lang),
            "cancellation_rate": format_percent(r.cancellation_rate, lang),
            "private_ratio": format_percent(r.private_ratio, lang),
        }
        for r in stats.rows
    ]

    return ReportPreview(
        report_key="instructor_workload",
        title=(
            "Eğitmen İş Yükü Raporu" if lang == "tr" else "Instructor Workload Report"
        ),
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("full_name", "Eğitmen", "Instructor", lang),
            _col("lesson_count", "Ders", "Lessons", lang),
            _col("total_hours", "Saat", "Hours", lang),
            _col("student_count", "Öğrenci", "Students", lang),
            _col("occupancy_rate", "Doluluk", "Occupancy", lang),
            _col("attendance_rate", "Devam", "Attendance", lang),
            _col("cancellation_rate", "İptal", "Cancellation", lang),
            _col("private_ratio", "Özel Ders %", "Private %", lang),
        ],
        rows=rows,
        totals={"total_hours": format_number(stats.total_hours, lang, 1)},
        summary={
            "disclaimer": stats.disclaimer_tr if lang == "tr" else stats.disclaimer_en,
            "average_occupancy": stats.average_occupancy,
        },
        row_count=len(rows),
    )


def _report_pool_usage(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    stats = pool_statistics(db, start, end, request.pool_id)

    rows = [
        {
            "label": d.label,
            "minutes": format_number(d.value, lang, 0),
            "hours": format_number(d.value / 60, lang, 1),
            "share": format_percent(d.percent, lang),
        }
        for d in stats.pool_usage
    ]
    rows.extend(
        {
            "label": f"{'Saat' if lang == 'tr' else 'Hour'} {p.label}",
            "minutes": format_number(p.value, lang, 0),
            "hours": format_number(p.value / 60, lang, 1),
            "share": "-",
        }
        for p in stats.hourly_load
        if p.value > 0
    )

    return ReportPreview(
        report_key="pool_usage",
        title="Havuz Kullanım Raporu" if lang == "tr" else "Pool Usage Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("label", "Havuz / Saat", "Pool / Hour", lang),
            _col("minutes", "Dakika", "Minutes", lang),
            _col("hours", "Saat", "Hours", lang),
            _col("share", "Pay", "Share", lang),
        ],
        rows=rows,
        totals={"overall_occupancy": format_percent(stats.overall_occupancy, lang)},
        summary={
            "busiest_hour": stats.busiest_hour,
            "quietest_hour": stats.quietest_hour,
            "most_used_lane": stats.most_used_lane,
            "free_capacity_hours": stats.free_capacity_hours,
        },
        row_count=len(rows),
    )


def _report_lane_occupancy(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    stats = pool_statistics(db, start, end, request.pool_id)

    rows = [
        {
            "lane": d.label,
            "minutes": format_number(d.value, lang, 0),
            "hours": format_number(d.value / 60, lang, 1),
            "share": format_percent(d.percent, lang),
        }
        for d in stats.lane_usage
    ]
    return ReportPreview(
        report_key="lane_occupancy",
        title="Kulvar Doluluk Raporu" if lang == "tr" else "Lane Occupancy Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("lane", "Kulvar", "Lane", lang),
            _col("minutes", "Dakika", "Minutes", lang),
            _col("hours", "Saat", "Hours", lang),
            _col("share", "Pay", "Share", lang),
        ],
        rows=rows,
        totals={"overall_occupancy": format_percent(stats.overall_occupancy, lang)},
        summary={"most_used_lane": stats.most_used_lane},
        row_count=len(rows),
    )


def _report_finance(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    currency = settings.app_currency

    income = (
        db.scalar(
            select(
                func.coalesce(func.sum(Payment.amount - Payment.refunded_amount), 0)
            ).where(
                Payment.payment_date >= start,
                Payment.payment_date <= end,
                Payment.status != PaymentStatus.CANCELLED,
            )
        )
        or 0
    )
    expense = (
        db.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date >= start, Expense.expense_date <= end
            )
        )
        or 0
    )

    method_rows = db.execute(
        select(
            Payment.method,
            func.coalesce(func.sum(Payment.amount - Payment.refunded_amount), 0),
        )
        .where(
            Payment.payment_date >= start,
            Payment.payment_date <= end,
            Payment.status != PaymentStatus.CANCELLED,
        )
        .group_by(Payment.method)
    ).all()
    category_rows = db.execute(
        select(Expense.category, func.coalesce(func.sum(Expense.amount), 0))
        .where(Expense.expense_date >= start, Expense.expense_date <= end)
        .group_by(Expense.category)
    ).all()

    rows = [
        {
            "item": "TOPLAM GELİR" if lang == "tr" else "TOTAL INCOME",
            "amount": format_currency(float(income), currency, lang),
            "type": "income",
        }
    ]
    rows.extend(
        {
            "item": f"  · {method}",
            "amount": format_currency(float(total), currency, lang),
            "type": "income",
        }
        for method, total in method_rows
    )
    rows.append(
        {
            "item": "TOPLAM GİDER" if lang == "tr" else "TOTAL EXPENSE",
            "amount": format_currency(float(expense), currency, lang),
            "type": "expense",
        }
    )
    rows.extend(
        {
            "item": f"  · {category}",
            "amount": format_currency(float(total), currency, lang),
            "type": "expense",
        }
        for category, total in category_rows
    )
    rows.append(
        {
            "item": "NET" if lang == "tr" else "NET",
            "amount": format_currency(float(income) - float(expense), currency, lang),
            "type": "net",
        }
    )

    return ReportPreview(
        report_key="finance",
        title="Finans Raporu" if lang == "tr" else "Finance Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("item", "Kalem", "Item", lang),
            _col("amount", "Tutar", "Amount", lang),
        ],
        rows=rows,
        totals={
            "income": round(float(income), 2),
            "expense": round(float(expense), 2),
            "net": round(float(income) - float(expense), 2),
        },
        summary={"currency": currency},
        row_count=len(rows),
    )


def _report_collections(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    payments = db.scalars(
        select(Payment)
        .where(
            Payment.payment_date >= start,
            Payment.payment_date <= end,
            Payment.status != PaymentStatus.CANCELLED,
        )
        .order_by(Payment.payment_date.desc())
    ).all()

    rows = [
        {
            "receipt": p.receipt_number,
            "date": format_date(p.payment_date, lang),
            "student": p.student.full_name if p.student else "-",
            "amount": format_currency(p.net_amount, p.currency, lang),
            "method": p.method,
            "description": p.description or "-",
        }
        for p in payments
    ]
    total = sum(p.net_amount for p in payments)

    return ReportPreview(
        report_key="collections",
        title="Tahsilat Raporu" if lang == "tr" else "Collections Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("receipt", "Fiş No", "Receipt", lang),
            _col("date", "Tarih", "Date", lang),
            _col("student", "Öğrenci", "Student", lang),
            _col("amount", "Tutar", "Amount", lang),
            _col("method", "Yöntem", "Method", lang),
            _col("description", "Açıklama", "Description", lang),
        ],
        rows=rows,
        totals={
            "total": format_currency(total, settings.app_currency, lang),
            "count": len(rows),
        },
        summary={"total_amount": round(total, 2)},
        row_count=len(rows),
    )


def _report_outstanding(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    invoices = db.scalars(
        select(Invoice)
        .where(Invoice.total_amount > Invoice.paid_amount)
        .order_by(Invoice.due_date)
    ).all()

    rows = [
        {
            "invoice": inv.invoice_number,
            "student": inv.student.full_name if inv.student else "-",
            "due_date": format_date(inv.due_date, lang),
            "total": format_currency(float(inv.total_amount), inv.currency, lang),
            "paid": format_currency(float(inv.paid_amount), inv.currency, lang),
            "balance": format_currency(inv.balance, inv.currency, lang),
            "days_overdue": inv.days_overdue,
        }
        for inv in invoices
    ]
    total_balance = sum(inv.balance for inv in invoices)

    return ReportPreview(
        report_key="outstanding",
        title="Bekleyen Alacak Raporu" if lang == "tr" else "Outstanding Receivables",
        generated_at=datetime.now(),
        period_label=format_date(date.today(), lang),
        columns=[
            _col("invoice", "Fatura", "Invoice", lang),
            _col("student", "Öğrenci", "Student", lang),
            _col("due_date", "Vade", "Due Date", lang),
            _col("total", "Toplam", "Total", lang),
            _col("paid", "Ödenen", "Paid", lang),
            _col("balance", "Kalan", "Balance", lang),
            _col("days_overdue", "Gecikme (gün)", "Days Overdue", lang),
        ],
        rows=rows,
        totals={"balance": format_currency(total_balance, settings.app_currency, lang)},
        summary={"count": len(rows), "total_balance": round(total_balance, 2)},
        row_count=len(rows),
    )


def _report_membership(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    stmt = select(Membership)
    if request.membership_status:
        stmt = stmt.where(Membership.status == request.membership_status)
    memberships = db.scalars(stmt.order_by(Membership.start_date.desc())).all()

    rows = [
        {
            "student": m.student.full_name if m.student else "-",
            "package": m.package.name if m.package else "-",
            "start_date": format_date(m.start_date, lang),
            "end_date": format_date(m.end_date, lang) if m.end_date else "-",
            "status": m.status,
            "credits": (
                f"{m.used_credits}/{m.total_credits}" if m.total_credits else "∞"
            ),
            "price": format_currency(float(m.price_paid), "TRY", lang),
        }
        for m in memberships
    ]

    return ReportPreview(
        report_key="membership",
        title="Üyelik Raporu" if lang == "tr" else "Membership Report",
        generated_at=datetime.now(),
        period_label=format_date(date.today(), lang),
        columns=[
            _col("student", "Öğrenci", "Student", lang),
            _col("package", "Paket", "Package", lang),
            _col("start_date", "Başlangıç", "Start", lang),
            _col("end_date", "Bitiş", "End", lang),
            _col("status", "Durum", "Status", lang),
            _col("credits", "Kullanım", "Usage", lang),
            _col("price", "Tutar", "Amount", lang),
        ],
        rows=rows,
        totals={"count": len(rows)},
        summary={
            "active": sum(
                1 for m in memberships if m.status == MembershipStatus.ACTIVE
            ),
            "expired": sum(
                1 for m in memberships if m.status == MembershipStatus.EXPIRED
            ),
            "frozen": sum(
                1 for m in memberships if m.status == MembershipStatus.FROZEN
            ),
        },
        row_count=len(rows),
    )


def _report_sales(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    memberships = db.scalars(
        select(Membership).where(
            Membership.start_date >= start, Membership.start_date <= end
        )
    ).all()

    per_package: dict[str, dict[str, float]] = {}
    for m in memberships:
        name = m.package.name if m.package else "-"
        bucket = per_package.setdefault(
            name, {"count": 0, "revenue": 0.0, "discount": 0.0}
        )
        bucket["count"] += 1
        bucket["revenue"] += float(m.price_paid)
        bucket["discount"] += float(m.discount_amount)

    rows = [
        {
            "package": name,
            "count": int(data["count"]),
            "revenue": format_currency(data["revenue"], settings.app_currency, lang),
            "discount": format_currency(data["discount"], settings.app_currency, lang),
            "average": format_currency(
                data["revenue"] / data["count"] if data["count"] else 0,
                settings.app_currency,
                lang,
            ),
        }
        for name, data in sorted(per_package.items(), key=lambda kv: -kv[1]["revenue"])
    ]
    total_revenue = sum(d["revenue"] for d in per_package.values())

    return ReportPreview(
        report_key="sales",
        title="Satış Raporu" if lang == "tr" else "Sales Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("package", "Paket", "Package", lang),
            _col("count", "Adet", "Count", lang),
            _col("revenue", "Gelir", "Revenue", lang),
            _col("discount", "İndirim", "Discount", lang),
            _col("average", "Ortalama", "Average", lang),
        ],
        rows=rows,
        totals={
            "revenue": format_currency(total_revenue, settings.app_currency, lang),
            "count": len(memberships),
        },
        summary={"total_revenue": round(total_revenue, 2)},
        row_count=len(rows),
    )


def _report_performance(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)

    if request.student_id:
        summary = student_performance_summary(db, request.student_id, start, end)
        rows = [
            {
                "event": f"{e.distance_m} m {e.stroke}",
                "records": e.record_count,
                "best": format_swim_time(e.best_time),
                "mean": format_swim_time(e.mean_time),
                "first": format_swim_time(e.first_time),
                "last": format_swim_time(e.last_time),
                "improvement": f"{e.improvement_seconds:+.2f} sn",
                "improvement_pct": format_percent(e.improvement_percent, lang, 2),
                "trend": e.trend,
            }
            for e in summary.events
        ]
        title_suffix = f" - {summary.student_name}"
        extra = summary.model_dump(mode="json", exclude={"events"})
    else:
        from app.services.statistics_engine import find_top_improvers

        improvers = find_top_improvers(
            db, lookback_days=(end - start).days + 1, limit=100
        )
        rows = [
            {
                "event": f"{r.student_name} · {r.distance_m} m {r.stroke}",
                "records": r.record_count,
                "best": format_swim_time(r.last_time),
                "mean": "-",
                "first": format_swim_time(r.first_time),
                "last": format_swim_time(r.last_time),
                "improvement": f"-{r.improvement_seconds:.2f} sn",
                "improvement_pct": format_percent(r.improvement_percent, lang, 2),
                "trend": "improving",
            }
            for r in improvers
        ]
        title_suffix = ""
        extra = {"top_improvers": len(rows)}

    return ReportPreview(
        report_key="performance",
        title=("Performans Raporu" if lang == "tr" else "Performance Report")
        + title_suffix,
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("event", "Etkinlik", "Event", lang),
            _col("records", "Kayıt", "Records", lang),
            _col("best", "En İyi", "Best", lang),
            _col("mean", "Ortalama", "Mean", lang),
            _col("first", "İlk", "First", lang),
            _col("last", "Son", "Last", lang),
            _col("improvement", "Gelişim", "Improvement", lang),
            _col("improvement_pct", "Gelişim %", "Improvement %", lang),
            _col("trend", "Eğilim", "Trend", lang),
        ],
        rows=rows,
        totals={"count": len(rows)},
        summary=extra,
        row_count=len(rows),
    )


def _report_competition(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    competitions = db.scalars(
        select(Competition).where(
            Competition.start_date >= start, Competition.start_date <= end
        )
    ).all()

    rows = []
    for comp in competitions:
        for event in comp.events:
            for entry in event.entries:
                if not entry.result_time_seconds:
                    continue
                rows.append(
                    {
                        "competition": comp.name,
                        "event": event.name,
                        "student": entry.student.full_name if entry.student else "-",
                        "time": format_swim_time(entry.result_time_seconds),
                        "rank": entry.rank or "-",
                        "medal": entry.medal or "-",
                        "pb": "✓" if entry.is_personal_best else "",
                        "record": "✓" if entry.is_club_record else "",
                    }
                )

    return ReportPreview(
        report_key="competition",
        title="Yarışma Raporu" if lang == "tr" else "Competition Report",
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("competition", "Yarışma", "Competition", lang),
            _col("event", "Etkinlik", "Event", lang),
            _col("student", "Sporcu", "Athlete", lang),
            _col("time", "Derece", "Time", lang),
            _col("rank", "Sıra", "Rank", lang),
            _col("medal", "Madalya", "Medal", lang),
            _col("pb", "KR", "PB", lang),
            _col("record", "Kulüp Rekoru", "Club Record", lang),
        ],
        rows=rows,
        totals={"count": len(rows)},
        summary={
            "competitions": len(competitions),
            "medals": sum(1 for r in rows if r["medal"] != "-"),
            "personal_bests": sum(1 for r in rows if r["pb"]),
        },
        row_count=len(rows),
    )


def _report_daily_manager(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    target = request.date_from or date.today()
    day_start = datetime.combine(target, datetime.min.time())
    day_end = datetime.combine(target, datetime.max.time())

    lessons = db.scalars(
        select(Lesson)
        .where(Lesson.start_at >= day_start, Lesson.start_at <= day_end)
        .order_by(Lesson.start_at)
    ).all()

    rows = []
    for lesson in lessons:
        attendances = db.scalars(
            select(Attendance).where(Attendance.lesson_id == lesson.id)
        ).all()
        present = sum(1 for a in attendances if a.status in PRESENT_STATUSES)
        rows.append(
            {
                "time": f"{lesson.start_at:%H:%M}-{lesson.end_at:%H:%M}",
                "lesson": lesson.title,
                "pool": lesson.pool.name if lesson.pool else "-",
                "lane": lesson.lane.display_name if lesson.lane else "-",
                "instructor": lesson.instructor.full_name if lesson.instructor else "-",
                "enrolled": lesson.enrolled_count,
                "present": present if attendances else "-",
                "status": lesson.status,
            }
        )

    collected = (
        db.scalar(
            select(
                func.coalesce(func.sum(Payment.amount - Payment.refunded_amount), 0)
            ).where(
                Payment.payment_date == target,
                Payment.status != PaymentStatus.CANCELLED,
            )
        )
        or 0
    )

    return ReportPreview(
        report_key="daily_manager",
        title="Günlük Yönetici Raporu" if lang == "tr" else "Daily Manager Report",
        generated_at=datetime.now(),
        period_label=format_date(target, lang),
        columns=[
            _col("time", "Saat", "Time", lang),
            _col("lesson", "Ders", "Lesson", lang),
            _col("pool", "Havuz", "Pool", lang),
            _col("lane", "Kulvar", "Lane", lang),
            _col("instructor", "Eğitmen", "Instructor", lang),
            _col("enrolled", "Kayıtlı", "Enrolled", lang),
            _col("present", "Gelen", "Present", lang),
            _col("status", "Durum", "Status", lang),
        ],
        rows=rows,
        totals={
            "lessons": len(rows),
            "collected": format_currency(float(collected), settings.app_currency, lang),
        },
        summary={
            "date": target.isoformat(),
            "collected_amount": round(float(collected), 2),
        },
        row_count=len(rows),
    )


def _report_management_summary(
    db: Session, request: ReportRequest, key: str
) -> ReportPreview:
    lang = request.language
    start, end = resolve_period(request.period, request.date_from, request.date_to)
    students = student_statistics(db, start, end)
    attendance = attendance_statistics(db, start, end)
    pools = pool_statistics(db, start, end)

    income = (
        db.scalar(
            select(
                func.coalesce(func.sum(Payment.amount - Payment.refunded_amount), 0)
            ).where(
                Payment.payment_date >= start,
                Payment.payment_date <= end,
                Payment.status != PaymentStatus.CANCELLED,
            )
        )
        or 0
    )
    expense = (
        db.scalar(
            select(func.coalesce(func.sum(Expense.amount), 0)).where(
                Expense.expense_date >= start, Expense.expense_date <= end
            )
        )
        or 0
    )

    def row(tr: str, en: str, value: Any) -> dict:  # noqa: ANN401
        return {"metric": tr if lang == "tr" else en, "value": value}

    rows = [
        row("Toplam Öğrenci", "Total Students", students.total_students),
        row("Aktif Öğrenci", "Active Students", students.active_students),
        row("Yeni Kayıt", "New Registrations", students.new_registrations),
        row("Ayrılan Öğrenci", "Lost Students", students.lost_students),
        row(
            "Tutundurma Oranı",
            "Retention Rate",
            format_percent(students.retention_rate, lang),
        ),
        row("Kayıp Oranı", "Churn Rate", format_percent(students.churn_rate, lang)),
        row(
            "Devam Oranı",
            "Attendance Rate",
            format_percent(attendance.overall_rate, lang),
        ),
        row(
            "Devamsızlık Oranı",
            "No-show Rate",
            format_percent(attendance.no_show_rate, lang),
        ),
        row(
            "Havuz Doluluk",
            "Pool Occupancy",
            format_percent(pools.overall_occupancy, lang),
        ),
        row("En Yoğun Saat", "Busiest Hour", pools.busiest_hour or "-"),
        row(
            "Toplam Gelir",
            "Total Income",
            format_currency(float(income), settings.app_currency, lang),
        ),
        row(
            "Toplam Gider",
            "Total Expense",
            format_currency(float(expense), settings.app_currency, lang),
        ),
        row(
            "Net",
            "Net",
            format_currency(
                float(income) - float(expense), settings.app_currency, lang
            ),
        ),
    ]

    titles = {
        "weekly_management": ("Haftalık Yönetim Raporu", "Weekly Management Report"),
        "monthly_management": ("Aylık Yönetim Raporu", "Monthly Management Report"),
    }
    title_tr, title_en = titles.get(key, ("Yönetim Raporu", "Management Report"))

    return ReportPreview(
        report_key=key,
        title=title_tr if lang == "tr" else title_en,
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("metric", "Gösterge", "Metric", lang),
            _col("value", "Değer", "Value", lang),
        ],
        rows=rows,
        totals={},
        summary={
            "income": round(float(income), 2),
            "expense": round(float(expense), 2),
            "net": round(float(income) - float(expense), 2),
        },
        row_count=len(rows),
    )


def _report_student_progress(db: Session, request: ReportRequest) -> ReportPreview:
    lang = request.language
    if not request.student_id:
        return ReportPreview(
            report_key="student_progress",
            title=(
                "Öğrenci Gelişim Raporu" if lang == "tr" else "Student Progress Report"
            ),
            generated_at=datetime.now(),
            period_label="-",
            columns=[
                _col("metric", "Gösterge", "Metric", lang),
                _col("value", "Değer", "Value", lang),
            ],
            rows=[],
            totals={},
            summary={"error": "student_id_required"},
            row_count=0,
        )

    start, end = resolve_period(request.period, request.date_from, request.date_to)
    student = db.get(Student, request.student_id)
    summary = student_performance_summary(db, request.student_id, start, end)

    attendance_rows = db.execute(
        select(Attendance.status, func.count(Attendance.id))
        .join(Lesson, Lesson.id == Attendance.lesson_id)
        .where(
            Attendance.student_id == request.student_id,
            Lesson.start_at >= datetime.combine(start, datetime.min.time()),
            Lesson.start_at <= datetime.combine(end, datetime.max.time()),
        )
        .group_by(Attendance.status)
    ).all()
    counts = dict(attendance_rows)
    total = sum(counts.values())
    present = sum(counts.get(s, 0) for s in PRESENT_STATUSES)

    rows = [
        {
            "metric": "Öğrenci" if lang == "tr" else "Student",
            "value": summary.student_name,
        },
        {
            "metric": "Seviye" if lang == "tr" else "Level",
            "value": student.swim_level if student else "-",
        },
        {
            "metric": "Devam Oranı" if lang == "tr" else "Attendance Rate",
            "value": format_percent(present / total * 100 if total else 0, lang),
        },
        {"metric": "Toplam Ders" if lang == "tr" else "Total Lessons", "value": total},
        {
            "metric": "Performans Kaydı" if lang == "tr" else "Performance Records",
            "value": summary.total_records,
        },
        {
            "metric": "Kişisel Rekor" if lang == "tr" else "Personal Bests",
            "value": summary.personal_best_count,
        },
        {
            "metric": "En Güçlü Stil" if lang == "tr" else "Strongest Stroke",
            "value": summary.strongest_stroke or "-",
        },
        {
            "metric": "Gelişim Alanı" if lang == "tr" else "Focus Area",
            "value": summary.weakest_stroke or "-",
        },
    ]
    rows.extend(
        {
            "metric": f"{e.distance_m} m {e.stroke}",
            "value": (
                f"{format_swim_time(e.best_time)} "
                f"({e.improvement_seconds:+.2f} sn, {e.trend})"
            ),
        }
        for e in summary.events
    )

    return ReportPreview(
        report_key="student_progress",
        title=(
            f"Öğrenci Gelişim Raporu - {summary.student_name}"
            if lang == "tr"
            else f"Student Progress Report - {summary.student_name}"
        ),
        generated_at=datetime.now(),
        period_label=f"{format_date(start, lang)} - {format_date(end, lang)}",
        columns=[
            _col("metric", "Gösterge", "Metric", lang),
            _col("value", "Değer", "Value", lang),
        ],
        rows=rows,
        totals={},
        summary=summary.model_dump(mode="json", exclude={"events"}),
        row_count=len(rows),
    )


REPORT_BUILDERS: dict[str, Callable[[Session, ReportRequest], ReportPreview]] = {
    "student_list": _report_student_list,
    "student_progress": _report_student_progress,
    "attendance": _report_attendance,
    "instructor_workload": _report_instructor_workload,
    "pool_usage": _report_pool_usage,
    "lane_occupancy": _report_lane_occupancy,
    "finance": _report_finance,
    "collections": _report_collections,
    "outstanding": _report_outstanding,
    "membership": _report_membership,
    "sales": _report_sales,
    "performance": _report_performance,
    "competition": _report_competition,
    "daily_manager": _report_daily_manager,
    "weekly_management": lambda db, req: _report_management_summary(
        db, req, "weekly_management"
    ),
    "monthly_management": lambda db, req: _report_management_summary(
        db, req, "monthly_management"
    ),
}


def build_report(db: Session, request: ReportRequest) -> ReportPreview:
    builder = REPORT_BUILDERS.get(request.report_key)
    if builder is None:
        from app.core.exceptions import NotFoundError

        raise NotFoundError(details={"report_key": request.report_key})
    return builder(db, request)


# ===========================================================================
# Dışa aktarma
# ===========================================================================
def to_csv(preview: ReportPreview) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([column["label"] for column in preview.columns])
    for row in preview.rows:
        writer.writerow([row.get(column["key"], "") for column in preview.columns])
    if preview.totals:
        writer.writerow([])
        for key, value in preview.totals.items():
            writer.writerow([key, value])
    # UTF-8 BOM: Excel'in Türkçe karakterleri doğru açması için
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_excel(preview: ReportPreview, org_name: str = "") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = preview.report_key[:31]

    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.cell(row=1, column=1, value=org_name or preview.title).font = title_font
    sheet.cell(row=2, column=1, value=preview.title)
    sheet.cell(row=3, column=1, value=f"{preview.period_label}")
    sheet.cell(row=4, column=1, value=f"{preview.generated_at:%d.%m.%Y %H:%M}")

    header_row = 6
    for index, column in enumerate(preview.columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column["label"])
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(preview.rows, start=header_row + 1):
        for column_index, column in enumerate(preview.columns, start=1):
            cell = sheet.cell(
                row=row_index, column=column_index, value=row.get(column["key"], "")
            )
            cell.border = border

    if preview.totals:
        total_row = header_row + len(preview.rows) + 2
        sheet.cell(row=total_row, column=1, value="TOPLAM / TOTAL").font = Font(
            bold=True
        )
        for offset, (key, value) in enumerate(preview.totals.items(), start=1):
            sheet.cell(row=total_row + offset, column=1, value=str(key))
            sheet.cell(row=total_row + offset, column=2, value=str(value)).font = Font(
                bold=True
            )

    for index, column in enumerate(preview.columns, start=1):
        width = (
            max(
                len(str(column["label"])),
                *(len(str(row.get(column["key"], ""))) for row in preview.rows[:200]),
            )
            if preview.rows
            else len(str(column["label"]))
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            45, max(12, width + 3)
        )

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def to_pdf(preview: ReportPreview, org_name: str = "") -> bytes:
    stream = io.BytesIO()
    wide = len(preview.columns) > 5
    document = SimpleDocTemplate(
        stream,
        pagesize=landscape(A4) if wide else A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=preview.title,
        author=org_name or settings.app_name,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleTR",
        parent=styles["Title"],
        fontName=_FONT_BOLD,
        fontSize=16,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "MetaTR",
        parent=styles["Normal"],
        fontName=_FONT_NAME,
        fontSize=9,
        textColor=colors.HexColor("#64748B"),
    )
    cell_style = ParagraphStyle(
        "CellTR", parent=styles["Normal"], fontName=_FONT_NAME, fontSize=8, leading=10
    )
    header_style = ParagraphStyle(
        "HeadTR",
        parent=styles["Normal"],
        fontName=_FONT_BOLD,
        fontSize=8.5,
        leading=10,
        textColor=colors.white,
    )

    story: list[Any] = [
        Paragraph(org_name or settings.app_name, meta_style),
        Paragraph(preview.title, title_style),
        Paragraph(
            f"Dönem / Period: {preview.period_label} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Oluşturma / Generated: {preview.generated_at:%d.%m.%Y %H:%M} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Kayıt / Rows: {preview.row_count}",
            meta_style,
        ),
        Spacer(1, 0.5 * cm),
    ]

    if preview.rows:
        data = [
            [Paragraph(column["label"], header_style) for column in preview.columns]
        ]
        for row in preview.rows[:1500]:
            data.append(
                [
                    Paragraph(str(row.get(column["key"], "")), cell_style)
                    for column in preview.columns
                ]
            )

        available = (landscape(A4)[0] if wide else A4[0]) - 3 * cm
        column_width = available / len(preview.columns)
        table = Table(
            data, colWidths=[column_width] * len(preview.columns), repeatRows=1
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0EA5E9")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F1F5F9")],
                    ),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

        if len(preview.rows) > 1500:
            story.append(Spacer(1, 0.3 * cm))
            story.append(
                Paragraph(
                    f"Not: {len(preview.rows)} satırın ilk 1500'ü gösterilmiştir. "
                    f"Tam veri için Excel/CSV dışa aktarımını kullanın.",
                    meta_style,
                )
            )
    else:
        story.append(
            Paragraph(
                "Seçilen kriterlere uygun veri bulunamadı. / No data found.", cell_style
            )
        )

    if preview.totals:
        story.append(Spacer(1, 0.6 * cm))
        totals_data = [
            [Paragraph("<b>TOPLAM / TOTAL</b>", cell_style), ""],
            *[
                [
                    Paragraph(str(key), cell_style),
                    Paragraph(f"<b>{value}</b>", cell_style),
                ]
                for key, value in preview.totals.items()
            ],
        ]
        totals_table = Table(totals_data, colWidths=[6 * cm, 6 * cm])
        totals_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
                ]
            )
        )
        story.append(totals_table)

    document.build(story)
    return stream.getvalue()


def export_report(db: Session, request: ReportRequest) -> tuple[bytes, str, str]:
    """(içerik, dosya_adı, mime_type) döndürür."""
    preview = build_report(db, request)
    org = _org_name(db)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base = f"{request.report_key}_{stamp}"

    if request.format == "csv":
        return to_csv(preview), f"{base}.csv", "text/csv; charset=utf-8"
    if request.format == "xlsx":
        return (
            to_excel(preview, org),
            f"{base}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if request.format == "pdf":
        return to_pdf(preview, org), f"{base}.pdf", "application/pdf"

    import json

    return (
        json.dumps(
            preview.model_dump(mode="json"), ensure_ascii=False, indent=2
        ).encode("utf-8"),
        f"{base}.json",
        "application/json",
    )
